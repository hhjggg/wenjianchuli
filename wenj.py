from openpyxl import load_workbook
import streamlit as st
import pandas as pd
from io import BytesIO

# ========== 网页界面 ==========
st.title("订单SN数据回填工具")
st.subheader("请上传两份Excel文件")

# 文件上传组件
sn_file = st.file_uploader("① 上传SN数据源表", type=["xlsx"])
target_file = st.file_uploader("② 上传待回填目标表", type=["xlsx"])

# 两个文件都上传后才执行处理
if sn_file is not None and target_file is not None:
    # 读取上传的表格
    df_sn = pd.read_excel(sn_file)
    df_target = pd.read_excel(target_file)
    st.success("文件读取完成，开始处理数据！")



    # 先pandas读入强制文本，再转成openpyxl工作簿
    temp_io = BytesIO()
    df_temp = pd.read_excel(sn_file, dtype=str)
    df_temp.to_excel(temp_io, index=False)
    temp_io.seek(0)
    wb_sn_source = load_workbook(temp_io)
    ws_sn_source = wb_sn_source.active
    sn_mapping = {}
    over_three_sn_order = []

    for row in range(2, ws_sn_source.max_row + 1):
        jd_order = ws_sn_source.cell(row=row, column=5).value
        sku_id = ws_sn_source.cell(row=row, column=3).value
        sn_code = ws_sn_source.cell(row=row, column=10).value

        match_key = str(jd_order).strip()
        if sn_code is not None and str(sn_code).strip() != "":
            if match_key not in sn_mapping:
                sn_mapping[match_key] = []
            sn_mapping[match_key].append((sku_id, sn_code))

    # 遍历字典，筛选有效SN数量>=3的订单
    for order, sku_sn_list in sn_mapping.items():
        if len(sku_sn_list) >= 3:
            over_three_sn_order.append(order)

    wb_sn_source.close()




    success_pair_list = []
    
    # 先用pandas强制全列文本读取，消除科学计数
    temp_buf = BytesIO()
    df_temp = pd.read_excel(target_file, dtype=str)
    df_temp.to_excel(temp_buf, index=False)
    temp_buf.seek(0)
    # 再交给openpyxl，你后面ws_target操作全部保留不变
    wb_target = load_workbook(temp_buf)
    ws_target = wb_target.active

    for row in range(2, ws_target.max_row + 1):
        tri_order = ws_target.cell(row=row, column=3).value
        ws_target.cell(row=row, column=7, value=None)
        ws_target.cell(row=row, column=8, value=None)
        if tri_order is None:
            continue
        current_key = str(tri_order).strip()
        if current_key not in sn_mapping:
            continue
        
        data_list = sn_mapping[current_key]
        success_pair_list.extend(data_list)
        sn_count = len(data_list)
        if sn_count == 1:
            _, sn = data_list[0]
            ws_target.cell(row=row, column=7, value=sn)
        elif sn_count == 2:
            data_sorted = sorted(data_list, key=lambda x: int(x[0]) if str(x[0]).isdigit() else 0)
            _, sn_small = data_sorted[0]
            _, sn_big = data_sorted[1]
            ws_target.cell(row=row, column=7, value=sn_big)
            ws_target.cell(row=row, column=8, value=sn_small)
        elif sn_count >= 3:
            ws_target.cell(row=row, column=7, value=sn_count)
            continue




        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_target.to_excel(writer, index=False)
        output.seek(0)

    # 网页下载按钮
    st.download_button(
        label="📥 下载处理完成Excel",
        data=output,
        file_name="回填完成_订单表.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("请依次上传两份xlsx格式Excel文件")